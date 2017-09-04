# -*- coding: utf-8 -*-
#Author: yexiao
#Date: 2017-09-03
#Purpose: 解析excel数据表到txt中，并生成对应的cpp文件

import os
import sys
import string
import xlrd
import json
import openpyxl
reload(sys)
sys.setdefaultencoding('utf8')

_errorDes = ["数据错误：","文件名","行数","列数"]
_isError = False

global excelPath
global txtPath


global isXls

# 按类型解析数据
dataType = ["int","bool","string","object","class"]
# cpp中对应的类型
cppType = ["int32","bool","FString","UObject*","UClass*"]
# 查找所有 xls
def findAllFile(callback):
    fileList = os.listdir(excelPath)
    for f in fileList:

        # 检查是否有错误
        if _isError:
            break

        filePath = os.path.join(excelPath,f)

        if f[0] == "." or f.find(".svn") > 0 or f.find(".DS_Store") > 0 or f.startswith("~$"):
            continue
    
        if os.path.isdir(filePath):
            findAllFile(filePath,callback)
        else:
            if f.endswith(".xls") or f.endswith(".xlsx"):
                callback(filePath,f)

# 解析 excel的type类型
# ctype类型 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
def parseTitleType(rowData,col):
    tableKey = []
    tableType=[]
    for c in range(col):
        typeValue = None
        keyValue = None
        if isXls:
            typeValue = rowData.cell(1,c).value
            keyValue = rowData.cell(2,c).value
        else:
            typeValue = rowData.cell(row=2,column=c+1).value
            keyValue = rowData.cell(row=3,column=c+1).value
        if typeValue == None:
            break
        tmp = typeValue.split('|')
        # 这里判断只导出客户端的
        if len(tmp) == 2 :
            if tmp[1]=='client':
                typeValue = tmp[0]
            else:
                typeValue = None
        tableKey.append(keyValue)
        tableType.append(typeValue)
    return tableKey,tableType

# 获取表格数据
def getDataByExcel(pTabData,pRow,tblKey):
    ret = []
    for row in range(pRow):
        coltab=[]
        for col in range(len(tblKey)):
            tmp = None
            global isXls
            if isXls:
                tmp = pTabData.cell(row+3,col).value
            else:
                tmp = pTabData.cell(row=row+4,column=col+1).value
            if tmp == None:
                tmp = ''
            coltab.append(tmp)
        ret.append(coltab)
    return ret

# 将excel行数据整理成txt
def parseData(pData,pType):
    ret = ""
    checkSame=set()
    # print pData
    # print pType
    for row in range(len(pData)):
        item=[]
        for col in range(len(pType)):
            tmp = pData[row][col]
            if col == 0:
                if tmp in checkSame:
                    print '存在唯一key:'+tmp
                    sys.exit(1)
                else:
                    checkSame.add(tmp)
            if tmp == None or tmp == '':
                tmp = ''
            if pType[col] == dataType[0]:
                if tmp != '':
                    tmp = int(tmp)
                else:
                    tmp = -1
            elif pType[col] == dataType[1]:
                if tmp == '0' or tmp == '':
                    tmp = '0'
                else:
                    tmp = '1'
            else:
                if tmp != '':
                    tmp= "\"%s\"" % (tmp.replace('\"','\\\"'))
                else:
                    tmp = "\"\""
            item.append(str(tmp))
        fileItem = ','.join(item) + ',\n'
        ret = ret+fileItem
    return ret

# 写文件
def writeFile(filePath,fileData):
    f = open(filePath,"w")
    f.write(fileData)
    f.close()
# 写文件
def writeToTxt(filePath,fileName,fileData):
    global excelPath
    global txtPath
    fDirPath = os.path.dirname(filePath)
    fDirPath = fDirPath.replace(excelPath,txtPath)
    if not os.path.exists(fDirPath):
        os.mkdir(fDirPath)
    filePath = os.path.join(fDirPath,"%s.txt" % (fileName))
    writeFile(filePath,fileData)
#h文件
h1 = '''
#pragma once
#include "Kismet/BlueprintFunctionLibrary.h"
#include "DB_%s.generated.h"

USTRUCT(BlueprintType)
struct F%s
{
    GENERATED_USTRUCT_BODY()
public:
    F%s(){};
    '''
h2 = '''
    UPROPERTY(EditAnywhere, BlueprintReadOnly, Category = "DB")
    %s %s;
'''
h3 = '''
};

UCLASS(Blueprintable)
class %s_API UDB_%s : public UBlueprintFunctionLibrary
{
    GENERATED_BODY()
public:

    UDB_%s();
    ~UDB_%s(){};
    bool loadData();

    UFUNCTION(BlueprintCallable, Category = "DB")
    static F%s get%sBy%s(%s _value);
    UFUNCTION(BlueprintCallable, Category = "DB")
    static TMap<%s,F%s> getAll%sDB();
};
'''

def writeToH(filePath,fileName,keyName,keyType):
    global excelPath
    global txtPath
    global projName
    cName = fileName.capitalize() #首字母大写
    uName = projName.upper() #转换为大写
    fDirPath = os.path.dirname(filePath)
    fDirPath = fDirPath.replace(excelPath,txtPath)
    if not os.path.exists(fDirPath):
        os.mkdir(fDirPath)
    filePath = os.path.join(fDirPath,"DB_%s.h" % (cName))
    a = h1%(cName,cName,cName)

    for i in range(len(keyType)):
        wType = keyType[i]
        for j in range(len(dataType)):
            dType = dataType[j]
            if wType == dType:
                a = a + (h2%(cppType[j],keyName[i]))
                if i == 0 :
                    key1 = keyName[i].capitalize()
                    ktype1 = cppType[j]
                break
    b = h3%(uName,cName,cName,cName,cName,cName,key1,ktype1,ktype1,cName,cName)

    writeFile(filePath,(a+b))
c1='''
#include "%s.h"
#include "DB_%s.h"

static TMap<%s,F%s> m_map;

UDB_Role::UDB_%s()
{
     loadData();
}
bool UDB_%s::loadData()
{
    m_map.Empty();
    FString path = FPaths::GameDir() + "Content/DB/DB_%s.txt";
    if (!FPlatformFileManager::Get().GetPlatformFile().FileExists(*path))
        return false;
    TArray<FString> db;
    FString contentStr;
    FFileHelper::LoadFileToString(contentStr,*path);
    contentStr.ParseIntoArray(db, TEXT("\\n"), false);
    for (int i = 0; i < db.Num(); i++)
    {
        FString aString = db[i];
        TArray<FString> array = {};
        aString.ParseIntoArray(array, TEXT(","), false);
        F%s dbS;
'''
c21='''
        dbS.%s = FCString::Atoi(*array[%s]);
'''
c22='''
        if (FCString::Atoi(*array[%s]) == 1)
            dbS.%s = true;
        else
            dbS.%s = false;
'''
c23='''
        dbS.%s = *array[%s];
'''
c24='''
        dbS.%s = LoadObject<UObject>(NULL, *array[%s]);
'''
c25='''
        dbS.%s = LoadClass<AActor>(NULL, *array[%s]);
'''
c3='''
        m_map.Add(dbS.%s, dbS);
'''
c4='''
    }
    return true;
}

FRole UDB_%s::get%sBy%s(%s _value);
{
    return m_map.FindRef(_value);
}
TMap<%s,F%s> UDB_%s::getAll%sDB()()
{
    return m_map
}
'''
# 写入cpp文件
def writeToCpp(filePath,fileName,keyName,keyType):
    global excelPath
    global txtPath
    global projName
    cName = fileName.capitalize() #首字母大写
    uName = projName.upper() #转换为大写
    fDirPath = os.path.dirname(filePath)
    fDirPath = fDirPath.replace(excelPath,txtPath)
    if not os.path.exists(fDirPath):
        os.mkdir(fDirPath)
    filePath = os.path.join(fDirPath,"DB_%s.cpp" % (cName))

    for j in range(len(dataType)):
        if dataType[j] == keyType[0]:
            ktype1 = cppType[j]
            break
    a = c1%(projName,cName,ktype1,cName,cName,cName,cName,cName)
    for i in range(len(keyType)):
        wType = keyType[i]
        for j in range(len(dataType)):
            dType = dataType[j]
            if wType == dType:
                if j == 0 :
                    a = a + (c21%(keyName[i],i))
                elif j == 1:
                    a = a + (c22%(i,keyName[i],keyName[i]))
                elif j == 2:
                    a = a + (c23%(keyName[i],i))
                elif j == 3:
                    a = a + (c24%(keyName[i],i))
                elif j == 4:
                    a = a + (c25%(keyName[i],i))
                if i == 0 :
                    key1 = keyName[i].capitalize()
                    if j == 0 or j == 2:
                        b = c3%(keyName[i])
                    else:
                        print 'first row must be int or string'
                        sys.exit(1)
                break
    a = a + b
    a = a + (c4%(cName,cName,key1,ktype1,ktype1,cName,cName,cName))
    writeFile(filePath,a)
#拆分excel
def parseExcel(filePath,fileName):
    global isXls
    if fileName.endswith(".xls"):
        isXls = True
        # 读取数据
        excel = xlrd.open_workbook(filePath)
        #获取workbook中所有的表格  
        sheets = excel.sheet_names() 
        #循环遍历所有sheet  
        for i in range(len(sheets)):
            sheet = excel.sheet_names()[i]
            pTabData = excel.sheet_by_index(i)
            if pTabData.nrows > 3 and pTabData.ncols > 0 :
                tabTitle = parseTitleType(pTabData,pTabData.ncols)
                data =getDataByExcel(pTabData,pTabData.nrows - 3,tabTitle[0])
                title = sheet.split('|')
                if len(title) == 2:
                    fName = "DB_%s" % (title[1].capitalize())
                    writeToTxt(filePath,fName,parseData(data,tabTitle[1])) #写入txt文件
                    # 写cpp 和 h 文件
                    writeToH(filePath,title[1],tabTitle[0],tabTitle[1])
                    writeToCpp(filePath,title[1],tabTitle[0],tabTitle[1])
                else:
                    print 'excel sheet配置错误'
                    sys.exit(1)
    else:
        isXls = False
        excel = openpyxl.load_workbook(filePath)
        sheets = excel.get_sheet_names() 
        #循环遍历所有sheet  
        for i in range(len(sheets)):  
            sheet= excel.get_sheet_by_name(sheets[i])
            if sheet.max_row > 3 and sheet.max_column > 0 :
                tabTitle = parseTitleType(sheet,sheet.max_column)
                data = getDataByExcel(sheet,sheet.max_row - 3,tabTitle[0])
                title = sheet.title.split('|')
                if len(title) == 2:
                    fName = "DB_%s" % (title[1].capitalize())
                    #写入txt文件
                    writeToTxt(filePath,fName,parseData(data,tabTitle[1])) 
                    # 写cpp 和 h 文件
                    writeToH(filePath,title[1],tabTitle[0],tabTitle[1])
                    writeToCpp(filePath,title[1],tabTitle[0],tabTitle[1])
                else:
                    print 'excel sheet配置错误'
                    sys.exit(1)



# 1.excel 文件夹路径 2.cpp、txt文件路径 3.工程名字
if __name__ == "__main__":
    global excelPath
    global txtPath
    global projName
    if len(sys.argv) <= 4:
        excelPath = sys.argv[1]
        if not os.path.exists(excelPath):
            print 'excelPath 路径不存在'
            sys.exit(1)
        txtPath = sys.argv[2]
        projName = sys.argv[3]
        # 检查并创建目录
        if not os.path.exists(txtPath):
            os.makedirs(txtPath)

        findAllFile(parseExcel)
        print 'parse excel success!'
    else:
        print '参数不正确'
        sys.exit(1)
